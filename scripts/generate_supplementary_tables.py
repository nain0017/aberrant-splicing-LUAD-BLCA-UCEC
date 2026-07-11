"""
step4_supplementary_tables_v2.py
Paper 5A — supplementary tables, single Excel workbook.

CHANGE FROM v1:
  - build_T8() now reads the single pre_commitments/STAGE_C_DEVIATIONS.md
    file (six H2 sections, one per deviation) instead of looking for
    multiple files under Stage2C_corrected/.
  - T8 row schema expanded: deviation_id, title, date, direction_of_change,
    rationale_excerpt, source_file.
  - D7 (dPSI_scn >= 0.05 floor) appended programmatically.

Writes:
  Stage2C_corrected/tables/Paper5A_SupplementaryTables.xlsx

Sheets:
  T1_LUAD, T1_BLCA, T1_UCEC   Primary aberrant catalog per cancer
  T2_LUAD, T2_BLCA, T2_UCEC   Secondary aberrant catalog per cancer
  T3_conserved_core           3-way conserved core (123 events, 115 genes)
  T4_top50_annotated          Top-50 conserved events with annotation
  T5_hallmark_matrix          Hallmark cancer-splicing genes x cancers
  T6_M4_audit                 6-test audit (locked from cell 122) + Woolf 95% CIs
  T7_pathway_enrichment       Conserved-core pathway enrichment
  T8_deviations               D1-D6 from STAGE_C_DEVIATIONS.md + D7
  T9_sample_manifest          Cohort-summary placeholder
  T10_software_versions       Locked tool versions and parameters
"""

import os
import sys
import glob
import re
from datetime import date

import numpy as np
import pandas as pd

try:
    import openpyxl  # noqa
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# ----------------------------------------------------------------------
# Config
# ----------------------------------------------------------------------
PROJECT_ROOT = "/content/drive/MyDrive/Paper5_SharedAntigens"
STAGE_C_DIR  = os.path.join(PROJECT_ROOT, "data/processed/psi/Stage2C_corrected")
OVERLAP_DIR  = os.path.join(STAGE_C_DIR, "M3prime_overlap")
M4_DIR       = os.path.join(STAGE_C_DIR, "M4_enrichment")
TABLE_DIR    = os.path.join(STAGE_C_DIR, "tables")
os.makedirs(TABLE_DIR, exist_ok=True)

CANCERS = ["LUAD", "BLCA", "UCEC"]

HALLMARK_GENES = [
    "FAS", "CD44", "FGFR1", "FGFR2", "FGFR3", "BCL2L1", "MDM4", "MDM2", "VEGFA",
    "KLF6", "PKM", "MET", "EGFR", "AR", "TP53", "RAC1", "RON", "SYK", "TIA1",
    "NUMB", "CTNNB1", "BIN1", "CASP9", "CASP8", "TRAF3", "ESR1", "BRCA1", "PIK3CA",
]


# ----------------------------------------------------------------------
# Helpers
# ----------------------------------------------------------------------
def load_corrected(c):
    path = os.path.join(STAGE_C_DIR, f"{c}_corrected_results.parquet")
    if not os.path.exists(path):
        sys.exit(f"Missing {path}")
    df = pd.read_parquet(path)
    if 'gene_name' in df.columns and 'gene_symbol' not in df.columns:
        df = df.rename(columns={'gene_name': 'gene_symbol'})
    return df


def standard_event_columns(df):
    cols = [
        'event_id', 'gene_symbol', 'gene_id', 'chrom', 'strand',
        'cassette_start', 'cassette_end',
        'n_tumor_well_covered', 'n_normal_well_covered', 'n_scn_well_covered',
        'psi_tumor_mean', 'psi_normal_mean', 'psi_scn_mean',
        'delta_psi', 'delta_psi_scn',
        'cohens_d',
        'bb_chi2', 'bb_pvalue', 'bb_mu_tumor', 'bb_mu_normal', 'bb_dispersion_phi',
        'mwu_u', 'mwu_pvalue',
        'q_value',
        'tier', 'has_scn_data', 'scn_concordant', 'scn_discordant',
        'both_tests_significant',
    ]
    cols = [c for c in cols if c in df.columns]
    return df[cols].copy()


def round_psi_columns(df):
    psi_cols = [c for c in df.columns
                if 'psi' in c.lower() or 'dispersion' in c.lower() or
                'cohens_d' == c or 'delta_psi' in c.lower()]
    for col in psi_cols:
        if col in df.columns:
            df[col] = df[col].astype(float).round(4)
    pcols = ['bb_pvalue', 'mwu_pvalue', 'q_value', 'bb_chi2']
    for col in pcols:
        if col in df.columns:
            df[col] = df[col].astype(float)
    return df


# ----------------------------------------------------------------------
# T1, T2: per-cancer event catalogs
# ----------------------------------------------------------------------
def build_T1_T2(cancer_dfs):
    sheets = {}
    for c in CANCERS:
        df = cancer_dfs[c]
        pri = df[df['tier'] == 'primary'].copy()
        sec = df[df['tier'] == 'secondary'].copy()
        pri = round_psi_columns(standard_event_columns(pri))
        sec = round_psi_columns(standard_event_columns(sec))
        pri = pri.reindex(pri['delta_psi'].abs().sort_values(ascending=False).index)
        sec = sec.reindex(sec['delta_psi'].abs().sort_values(ascending=False).index)
        sheets[f"T1_{c}"] = pri
        sheets[f"T2_{c}"] = sec
        print(f"  T1_{c}: {len(pri):,} primary | T2_{c}: {len(sec):,} secondary")
    return sheets


# ----------------------------------------------------------------------
# T3: 3-way conserved core
# ----------------------------------------------------------------------
def build_T3(cancer_dfs):
    path = os.path.join(OVERLAP_DIR, "three_way_shared_events.parquet")
    if not os.path.exists(path):
        print("  [T3] three_way_shared_events.parquet not found; building from primaries")
        pri_ids = [set(cancer_dfs[c][cancer_dfs[c]['tier'] == 'primary']['event_id'].astype(str))
                   for c in CANCERS]
        common = pri_ids[0] & pri_ids[1] & pri_ids[2]
        rows = []
        for eid in sorted(common):
            row = {'event_id': eid}
            for c in CANCERS:
                hit = cancer_dfs[c][cancer_dfs[c]['event_id'].astype(str) == eid]
                if len(hit) > 0:
                    h = hit.iloc[0]
                    row[f'gene_symbol_{c}']    = h.get('gene_symbol', '')
                    row[f'delta_psi_{c}']      = round(float(h['delta_psi']), 4)
                    row[f'q_value_{c}']        = float(h['q_value'])
                    row[f'cohens_d_{c}']       = round(float(h['cohens_d']), 3)
            rows.append(row)
        df = pd.DataFrame(rows)
    else:
        df = pd.read_parquet(path)
        for col in df.select_dtypes(include=[np.number]).columns:
            if 'psi' in col.lower() or 'cohens' in col.lower():
                df[col] = df[col].round(4)
    print(f"  T3_conserved_core: {len(df):,} rows")
    return df


# ----------------------------------------------------------------------
# T4: top-50 annotated
# ----------------------------------------------------------------------
def build_T4():
    path = os.path.join(OVERLAP_DIR, "top50_conserved_events_annotated.csv")
    if not os.path.exists(path):
        print("  [T4] top50_conserved_events_annotated.csv not found; skipping")
        return None
    df = pd.read_csv(path)
    print(f"  T4_top50_annotated: {len(df):,} rows")
    return df


# ----------------------------------------------------------------------
# T5: hallmark matrix
# ----------------------------------------------------------------------
def build_T5(cancer_dfs):
    rows = []
    for g in HALLMARK_GENES:
        row = {'gene_symbol': g}
        any_hit = False
        for c in CANCERS:
            pri = cancer_dfs[c][(cancer_dfs[c]['tier'] == 'primary') &
                                (cancer_dfs[c]['gene_symbol'] == g)]
            if len(pri) == 0:
                row[f'in_primary_{c}'] = False
                row[f'strongest_dpsi_{c}'] = np.nan
                row[f'n_events_{c}'] = 0
            else:
                any_hit = True
                pri = pri.copy()
                pri['abs_dpsi'] = pri['delta_psi'].abs()
                best = pri.sort_values('abs_dpsi', ascending=False).iloc[0]
                row[f'in_primary_{c}'] = True
                row[f'strongest_dpsi_{c}'] = round(float(best['delta_psi']), 4)
                row[f'n_events_{c}'] = int(len(pri))
        n_cancers = sum(1 for c in CANCERS if row[f'in_primary_{c}'])
        row['n_cancers_hit'] = n_cancers
        if any_hit:
            rows.append(row)
    df = pd.DataFrame(rows)
    df = df.sort_values(['n_cancers_hit', 'gene_symbol'],
                        ascending=[False, True]).reset_index(drop=True)
    print(f"  T5_hallmark_matrix: {len(df):,} rows ({len(df)} of {len(HALLMARK_GENES)} hallmarks hit)")
    return df


# ----------------------------------------------------------------------
# T6: M4 audit
# ----------------------------------------------------------------------
def build_T6():
    BACKGROUND = 19782
    ABERRANT = 940
    TESTS = [
        ("Test 1",  "71-gene curated",                 "baseline",     71,  29, 14.25, 2.83e-20, "PASS", "OR>5, p<0.01"),
        ("Test 2",  "MSigDB EMT (n=200)",              "independent", 200,  22,  2.51, 2.22e-04, "FAIL", "OR>3, p<0.01"),
        ("Test 3",  "MSigDB Myogenesis (n=200)",       "independent", 200,  27,  3.19, 9.86e-07, "PASS", "OR>3, p<0.01"),
        ("Test 4a", "KRAS Signaling Up",               "neg_control", 200,  10,  1.06, 4.81e-01, "n/a",  "neg control"),
        ("Test 4b", "DNA Repair",                      "neg_control", 150,   8,  1.13, 4.21e-01, "n/a",  "neg control"),
        ("Test 4c", "Allograft Rejection",             "neg_control", 200,  15,  1.64, 5.47e-02, "n/a",  "neg control"),
        ("Test 4d", "Pancreas Beta Cells",             "neg_control",  40,   1,  0.51, 8.58e-01, "n/a",  "neg control"),
        ("Test 4e", "Heme Metabolism",                 "neg_control", 200,  13,  1.40, 1.58e-01, "n/a",  "neg control"),
        ("Test 5",  "Randomization (1,000 random sets)","independent",   0,   0, np.nan, 0.001,  "PASS", "empirical p<0.01"),
        ("Test 6",  "De-circularized (n=54)",          "defensive",    54,  12,  5.79, 6.68e-06, "PASS", "OR>3, p<0.01"),
    ]
    rows = []
    for test_id, ref_name, group, ref_size, overlap, or_rep, p_rep, verdict, criterion in TESTS:
        row = {
            'test_id':           test_id,
            'reference_set':     ref_name,
            'group':             group,
            'reference_size':    ref_size if ref_size > 0 else None,
            'overlap':           overlap if overlap > 0 else None,
            'aberrant_union':    ABERRANT,
            'background':        BACKGROUND,
            'odds_ratio':        round(or_rep, 3) if not np.isnan(or_rep) else None,
            'p_value':           p_rep,
            'verdict':           verdict,
            'audit_criterion':   criterion,
        }
        if ref_size > 0 and overlap > 0:
            a, b, c, d = overlap, ABERRANT - overlap, ref_size - overlap, BACKGROUND - ABERRANT - (ref_size - overlap)
            if all(x > 0 for x in [a, b, c, d]):
                OR = (a * d) / (b * c)
                se = np.sqrt(1/a + 1/b + 1/c + 1/d)
                lo = float(np.exp(np.log(OR) - 1.96 * se))
                hi = float(np.exp(np.log(OR) + 1.96 * se))
                row['ci95_lower'] = round(lo, 3)
                row['ci95_upper'] = round(hi, 3)
        rows.append(row)
    df = pd.DataFrame(rows)
    print(f"  T6_M4_audit: {len(df):,} rows (locked from cell 122)")
    return df


# ----------------------------------------------------------------------
# T7: pathway enrichment
# ----------------------------------------------------------------------
def build_T7():
    patterns = ['pathway*.csv', '*enrichr*.csv', '*enrich*.csv',
                'gseapy*.csv', '*hallmark*.csv', '*GO_*.csv', '*reactome*.csv']
    found = []
    for p in patterns:
        found.extend(glob.glob(os.path.join(OVERLAP_DIR, p)))
    found = sorted(set(found))
    if not found:
        print("  [T7] no pathway enrichment CSVs found in M3prime_overlap/; skipping")
        return None
    print(f"  [T7] concatenating {len(found)} pathway CSVs:")
    frames = []
    for f in found:
        try:
            sub = pd.read_csv(f)
            sub['source_file'] = os.path.basename(f)
            frames.append(sub)
            print(f"    + {os.path.basename(f)}  ({len(sub):,} rows)")
        except Exception as e:
            print(f"    ! {os.path.basename(f)}: {e}")
    if not frames:
        return None
    df = pd.concat(frames, ignore_index=True, sort=False)
    print(f"  T7_pathway_enrichment: {len(df):,} rows total")
    return df


# ----------------------------------------------------------------------
# T8: deviations register  --- PATCHED ---
# ----------------------------------------------------------------------
def build_T8():
    """Parse pre_commitments/STAGE_C_DEVIATIONS.md as a single file with
    H2 deviation sections, extract one row per deviation, then append D7."""
    md_path = os.path.join(PROJECT_ROOT, "pre_commitments/STAGE_C_DEVIATIONS.md")
    rows = []
    if not os.path.exists(md_path):
        print(f"  [T8] {md_path} not found; writing only D7")
    else:
        with open(md_path, 'r') as f:
            txt = f.read()
        print(f"  [T8] parsing {md_path}  ({len(txt):,} chars)")

        # Split on H2 headers that look like "## Deviation N — Title".
        # Lookahead so the delimiter is kept with the next section.
        sections = re.split(r'(?m)^(?=## Deviation \d+)', txt)
        for sec in sections:
            sec = sec.strip()
            if not sec.startswith('## Deviation'):
                continue
            lines = sec.splitlines()
            header = lines[0].lstrip('#').strip()
            # Match either em-dash, en-dash, or hyphen between number and title
            m = re.match(r'Deviation\s+(\d+)\s*[—–-]\s*(.+)', header)
            if not m:
                continue
            num = int(m.group(1))
            title = m.group(2).strip()

            # Pull the **Date:** / **Direction:** / **Author:** lines if nearby
            dev_date = ""
            direction = ""
            for line in lines[1:8]:
                if line.startswith('**Date:**'):
                    dev_date = line.replace('**Date:**', '').strip()
                elif line.startswith('**Direction of change:**'):
                    direction = line.replace('**Direction of change:**', '').strip()

            # Find a substantive section as the rationale excerpt
            rationale = ""
            for h_label in ['### What changed', '### What happened',
                            '### What was pre-committed', '### Why this record exists',
                            '### Numbers locked']:
                idx = sec.find(h_label)
                if idx != -1:
                    after = sec[idx + len(h_label):].lstrip()
                    next_h = re.search(r'(?m)^###\s', after)
                    snippet = after[:next_h.start()] if next_h else after
                    snippet = re.sub(r'\s+', ' ', snippet).strip()
                    rationale = snippet[:600] + ('...' if len(snippet) > 600 else '')
                    break

            rows.append({
                'deviation_id':         f"D{num}",
                'title':                title,
                'date':                 dev_date,
                'direction_of_change':  direction,
                'rationale_excerpt':    rationale,
                'source_file':          'pre_commitments/STAGE_C_DEVIATIONS.md',
            })
            print(f"    + D{num}: {title[:65]}")

    # Append D7 (dPSI_scn >= 0.05 floor)
    rows.append({
        'deviation_id':         'D7',
        'title':                'dPSI_scn >= 0.05 floor in 3C concordance check',
        'date':                 date.today().isoformat(),
        'direction_of_change':  'Strengthening (more conservative).',
        'rationale_excerpt':    (
            'Stage 2C corrected pipeline (cell 107) requires not only sign(dPSI_tumor_vs_normal) '
            '== sign(dPSI_tumor_vs_SCN) but also |dPSI_scn| >= 0.05, to exclude events where the '
            'SCN-direction call is too weak to be informative. This tightens the primary tier '
            'and is conservative. Logged as a post-hoc precommit clarification identified during '
            'figure regeneration audit (2026-06-13).'),
        'source_file':          '(this audit pass; to be appended to STAGE_C_DEVIATIONS.md)',
    })

    df = pd.DataFrame(rows)
    print(f"  T8_deviations_register: {len(df):,} rows")
    return df


# ----------------------------------------------------------------------
# T9: sample manifest (best-effort)
# ----------------------------------------------------------------------
def build_T9():
    candidates = [
        os.path.join(PROJECT_ROOT, "data/processed/clinical/sample_manifest.csv"),
        os.path.join(PROJECT_ROOT, "data/processed/clinical/sample_manifest.tsv"),
        os.path.join(PROJECT_ROOT, "data/processed/sample_manifest.csv"),
        os.path.join(PROJECT_ROOT, "data/raw/recount3_metadata.csv"),
    ]
    for p in candidates:
        if os.path.exists(p):
            try:
                sep = '\t' if p.endswith('.tsv') else ','
                df = pd.read_csv(p, sep=sep)
                print(f"  T9_sample_manifest: loaded from {os.path.basename(p)} "
                      f"({len(df):,} rows)")
                return df
            except Exception as e:
                print(f"  T9 [{p}]: {e}")
    print("  [T9] no sample manifest found; writing summary placeholder.")
    placeholder = pd.DataFrame([
        {'cohort': 'LUAD',     'source': 'TCGA',     'n_samples': 542, 'recount3_project': 'TCGA',  'notes': 'cancer cohort'},
        {'cohort': 'LUNG',     'source': 'GTEx',     'n_samples': 655, 'recount3_project': 'GTEX',  'notes': 'normal cohort for LUAD'},
        {'cohort': 'LUAD-SCN', 'source': 'TCGA',     'n_samples': 59,  'recount3_project': 'TCGA',  'notes': 'solid-tissue normal for LUAD 3C check'},
        {'cohort': 'BLCA',     'source': 'TCGA',     'n_samples': 414, 'recount3_project': 'TCGA',  'notes': 'cancer cohort'},
        {'cohort': 'BLADDER',  'source': 'GTEx',     'n_samples': 21,  'recount3_project': 'GTEX',  'notes': 'normal cohort for BLCA'},
        {'cohort': 'BLCA-SCN', 'source': 'TCGA',     'n_samples': 19,  'recount3_project': 'TCGA',  'notes': 'solid-tissue normal for BLCA 3C check'},
        {'cohort': 'UCEC',     'source': 'TCGA',     'n_samples': 554, 'recount3_project': 'TCGA',  'notes': 'cancer cohort'},
        {'cohort': 'UTERUS',   'source': 'GTEx',     'n_samples': 159, 'recount3_project': 'GTEX',  'notes': 'normal cohort for UCEC'},
        {'cohort': 'UCEC-SCN', 'source': 'TCGA',     'n_samples': 35,  'recount3_project': 'TCGA',  'notes': 'solid-tissue normal for UCEC 3C check'},
    ])
    print(f"  T9_sample_manifest: {len(placeholder):,} rows (cohort-summary placeholder)")
    return placeholder


# ----------------------------------------------------------------------
# T10: software versions
# ----------------------------------------------------------------------
def build_T10():
    rows = [
        ('Python',              '3.12.13',                                 'language',          'Colab'),
        ('NumPy',               'see env',                                 'numerics',          'rint-based integer count reconstruction'),
        ('Pandas',              'see env',                                 'data frames',       'parquet I/O'),
        ('SciPy',               'see env',                                 'statistics',        'chi2 LRT, Mann-Whitney U, Nelder-Mead optimizer'),
        ('statsmodels',         'see env',                                 'multiple testing',  'BH-FDR (multipletests, method=fdr_bh)'),
        ('matplotlib',          'see env',                                 'figures',           'fonttype=42 for vector PDF/Illustrator'),
        ('matplotlib-venn',     'see env',                                 'figures',           'F4 Venn'),
        ('adjustText',          'optional',                                'figures',           'F3 hallmark labels (rotated 90 if missing)'),
        ('openpyxl',            'see env',                                 'spreadsheet',       'this workbook'),
        ('gseapy',              'see env',                                 'enrichment',        'pathway enrichment of conserved core'),
        ('recount3',            '(Bioconductor R)',                        'data',              'TCGA + GTEx Monorail-aligned BAMs and junctions'),
        ('SUPPA2',              '(parsed events)',                         'data',              'cassette exon event catalog (suppa_events_parsed.parquet, 38,951 events)'),
        ('GENCODE',             'v26',                                     'annotation',        'gene/transcript coordinates'),
        ('MyGene.info',         'API',                                     'ID resolution',     'Ensembl gene ID -> HGNC symbol; 828/844 resolved (98.1%)'),
        ('MSigDB',              'hallmark gene sets',                      'enrichment',        'EMT, Myogenesis, neg-control hallmarks'),
        ('Test (primary)',      'beta-binomial LRT',                       'parameter',         'logit link; shared dispersion phi; chi-sq df=1'),
        ('Test (sensitivity)',  'Mann-Whitney rank-sum (two-sided)',       'parameter',         'scipy.stats.mannwhitneyu, nan_policy=omit'),
        ('FDR',                 'Benjamini-Hochberg',                      'parameter',         'q < 0.05'),
        ('Effect size gate',    '|dPSI| >= 0.15 (mean-based)',             'parameter',         'replaces median to avoid snapping'),
        ('Coverage gate',       '>= 10 reads/sample, >= 20 samples/side',  'parameter',         'Kahles 2018 + Buen Abad Najar 2020'),
        ('3C concordance',      'sign(dPSI) == sign(dPSI_SCN) AND |dPSI_SCN| >= 0.05',
         'parameter',
         'SCN well-covered count >= 5; floor 0.05 logged as Deviation 7'),
        ('M2 prime thresholds', 'LUAD >= 500, BLCA >= 300, UCEC >= 300',   'parameter',         'precommit STAGE_C_FIX_PRECOMMIT.md Section 3'),
        ('Background size',     '19,782 protein-coding genes',             'parameter',         'M4 enrichment universe (MyGene-resolved)'),
        ('Random seed',         '42',                                      'parameter',         'F5 jitter; S2 calibration permutation; bootstrap'),
        ('Bootstrap',           '200 reps, 10% sample dropout',            'parameter',         'directional sign-stability on conserved core'),
        ('Calibration',         'label-permutation, 2,000 events/cancer',  'parameter',         'S2 lambda = 1.067 (LUAD) / 0.904 (BLCA) / 0.943 (UCEC)'),
    ]
    df = pd.DataFrame(rows, columns=['tool_or_parameter', 'version_or_value', 'role', 'notes'])
    print(f"  T10_software_versions: {len(df):,} rows")
    return df


# ----------------------------------------------------------------------
# Excel writer
# ----------------------------------------------------------------------
def write_workbook(sheets_in_order, out_path):
    if not HAS_OPENPYXL:
        sys.exit("openpyxl not available. Run:  !pip install openpyxl -q  and rerun.")
    with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
        for name, df in sheets_in_order:
            if df is None or len(df) == 0:
                print(f"  [skip] {name} is empty")
                continue
            sheet = name[:31]
            df.to_excel(writer, sheet_name=sheet, index=False, freeze_panes=(1, 0))

        wb = writer.book
        header_font = Font(bold=True, color="000000")
        header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        thin = Side(border_style="thin", color="BBBBBB")
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='left', vertical='center')
                cell.border = Border(bottom=thin)
            for col_idx, col_cells in enumerate(ws.columns, start=1):
                lengths = [len(str(c.value)) if c.value is not None else 0
                           for c in col_cells[:60]]
                width = min(max(max(lengths) + 2 if lengths else 12, 10), 42)
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width


# ----------------------------------------------------------------------
# Main
# ----------------------------------------------------------------------
def main():
    if not os.path.exists("/content/drive/MyDrive"):
        sys.exit("Drive not mounted.")
    if not os.path.exists(STAGE_C_DIR):
        sys.exit(f"Stage2C_corrected not found: {STAGE_C_DIR}")
    if not HAS_OPENPYXL:
        sys.exit("openpyxl missing. Run:  !pip install openpyxl -q  and rerun.")

    print(f"Project root: {PROJECT_ROOT}")
    print(f"Output dir:   {TABLE_DIR}")
    print(f"Date:         {date.today().isoformat()}")
    print()
    print("Loading corrected results parquets ...")
    cancer_dfs = {c: load_corrected(c) for c in CANCERS}

    print()
    print("Building T1/T2 per-cancer catalogs ...")
    t1_t2 = build_T1_T2(cancer_dfs)

    print()
    print("Building T3 conserved core ...")
    t3 = build_T3(cancer_dfs)

    print()
    print("Building T4 top-50 annotated ...")
    t4 = build_T4()

    print()
    print("Building T5 hallmark matrix ...")
    t5 = build_T5(cancer_dfs)

    print()
    print("Building T6 M4 audit ...")
    t6 = build_T6()

    print()
    print("Building T7 pathway enrichment ...")
    t7 = build_T7()

    print()
    print("Building T8 deviations register ...")
    t8 = build_T8()

    print()
    print("Building T9 sample manifest ...")
    t9 = build_T9()

    print()
    print("Building T10 software versions ...")
    t10 = build_T10()

    sheets_in_order = []
    for c in CANCERS:
        sheets_in_order.append((f"T1_{c}_primary", t1_t2.get(f"T1_{c}")))
    for c in CANCERS:
        sheets_in_order.append((f"T2_{c}_secondary", t1_t2.get(f"T2_{c}")))
    sheets_in_order += [
        ("T3_conserved_core",    t3),
        ("T4_top50_annotated",   t4),
        ("T5_hallmark_matrix",   t5),
        ("T6_M4_audit",          t6),
        ("T7_pathway_enrichment", t7),
        ("T8_deviations",        t8),
        ("T9_sample_manifest",   t9),
        ("T10_software",         t10),
    ]

    out_path = os.path.join(TABLE_DIR, "Paper5A_SupplementaryTables.xlsx")
    print()
    print(f"Writing {out_path} ...")
    write_workbook(sheets_in_order, out_path)
    size_kb = os.path.getsize(out_path) / 1024
    print(f"  Wrote {size_kb:.1f} KB")
    print()
    print("Sheet summary:")
    for name, df in sheets_in_order:
        n = 0 if df is None else len(df)
        ncols = 0 if df is None else len(df.columns)
        print(f"  {name:<28} {n:>6,} rows x {ncols:>3} cols")
    print()
    print("Supplementary tables done. Open the workbook in Excel/LibreOffice and skim.")


if __name__ == "__main__":
    main()